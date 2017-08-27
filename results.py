import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import openpyxl
 
def init_driver():
    driver = webdriver.Firefox()
    driver.wait = WebDriverWait(driver, 5)
    return driver


def login(driver,roll):

    try:
        box1 = driver.wait.until(EC.presence_of_element_located((By.NAME, "regno")))
        box2 = driver.wait.until(EC.presence_of_element_located((By.NAME, "schcode")))
        box1.send_keys(roll)#enter roll
        box2.send_keys('*****')#enter p/w
        btn= driver.wait.until(EC.presence_of_element_located((By.NAME, "B1")))
        btn.click()            
        # button.click()
    except TimeoutException:
        print("Box or Button not found in moodle.com")


if __name__ == "__main__":
    driver = init_driver()
    wb=openpyxl.Workbook()
    sheet=wb.create_sheet(title="Results")
    j=1
    for i in range(*******,*******):
        driver.get("http://resultsarchives.nic.in/cbseresults/cbseresults2016/class12/cbse1216revised.htm")    
        login(driver,i)
        data=driver.page_source
        soup=BeautifulSoup(data,"html.parser")


        namet=soup.find_all('tbody')[4].find_all('tr')[1].find_all('td')[1]
        table=soup.find_all('tbody')[5]
        sheet.cell(column=1,row=j,value=namet.text)
        sub=table.find_all('tr')
        s=0
        for i in range(1,6):
            t = sub[i].find_all('td')[4].text
            sheet.cell(column=i+1, row=j,value=t)
            s=s+int(t)
        sheet.cell(column=7, row=j,value= s/5.00)
        j=j+1
    time.sleep(1)
    wb.save("CBSE.xlsx")
    driver.quit()
